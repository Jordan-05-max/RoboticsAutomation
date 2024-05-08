import os
import time
import os.path
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename

Tk().withdraw()
file = askopenfilename()
loc = r"C:/Users/jorda/Documents/path.xlsx"

# To open the Workbook
wb_obj = openpyxl.load_workbook(loc)
sheet_obj = wb_obj.active
# For row 0 and column 0
cell_obj = sheet_obj.cell(row=2, column=1)
wb_obj.save(loc)
var = cell_obj.value
wb_obj.save(loc)

# print(var)


def fileName():
    print("File name with extention: "+os.path.basename(var))
    # return os.path.basename(var)




def fileWithout_ext():
    # print("File name without extension: " + os.path.splitext(os.path.basename(var))[0])
    return os.path.splitext(os.path.basename(var))[0]




def filesize():
    file_stats = os.stat(var)
    return f'File Size in Bytes is: {file_stats.st_size}'




def filesizeMB():
    file_stats = os.stat(var)

    return f'File Size in MegaBytes is: {file_stats.st_size / (1024 * 1024)}'




def fileCreation():
   return "Created: %s" % time.ctime(os.path.getctime(var))




def filemodification_info():
    return "Last modified: %s" % time.ctime(os.path.getmtime(var))


fileName()
fileWithout_ext()
filesize()
filesizeMB()
fileCreation()
filemodification_info()


# path = 'C:/Users/jorda/Documents/fileinfo.xlsx'
# wb_object = openpyxl.load_workbook(path)
#
# sheet_object = wb_object.active
#
#
# _var1 = sheet_object['A2']
# wb_obj.save(path)
# _var1.value = fileName()
# wb_object.save(path)
#
# _var2 = sheet_object['B2']
# _var2.value = fileWithout_ext()
# wb_object.save(path)
#
# _var3 = sheet_object['C2']
# _var3.value = filesize()
# wb_object.save(path)
#
# _var4 = sheet_object['D2']
# _var4.value = filesizeMB()
# wb_object.save(path)
#
# _var5 = sheet_object['E2']
# _var5.value = fileCreation()
# wb_object.save(path)
#
# _var6 = sheet_object['F2']
# _var6.value = filemodification_info()
# wb_object.save(path)
